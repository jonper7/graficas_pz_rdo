import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import matplotlib.pyplot as plt

def guardar_graficos_en_lote(graficos_info, excel_path):
    """
    Inserta múltiples gráficos en Excel (mucho más rápido).
    
    Parámetros:
        graficos_info : Lista de tuplas (png_path, sheet_name, cell, inst_name)
        excel_path    : Ruta del archivo Excel
    """
    
    print("\n" + "="*50)
    print("📝 Insertando gráficos en Excel...")
    print("="*50)
    
    try:
        wb = load_workbook(excel_path)
        insertados = 0
        omitidos = 0
        
        for png_path, sheet_name, cell, inst_name in graficos_info:
            try:
                # Verificar si la hoja existe
                if sheet_name not in wb.sheetnames:
                    print(f"  ⚠️ Hoja '{sheet_name}' no existe. Omitiendo gráfico de {inst_name}")
                    omitidos += 1
                    continue
                
                ws = wb[sheet_name]
                
                # ELIMINAR solo imágenes en la celda específica (si existen)
                col_idx = ord(cell[0].upper()) - ord('A')
                row_idx = int(''.join(filter(str.isdigit, cell))) - 1
                
                imagenes_a_eliminar = []
                for img in ws._images:
                    try:
                        # Intentar acceder al anchor de forma segura
                        if hasattr(img, 'anchor') and hasattr(img.anchor, '_from'):
                            # Si tiene el formato correcto, verificar posición
                            if (hasattr(img.anchor._from, 'col') and 
                                hasattr(img.anchor._from, 'row') and
                                img.anchor._from.col == col_idx and 
                                img.anchor._from.row == row_idx):
                                imagenes_a_eliminar.append(img)
                        else:
                            # Si no tiene el formato esperado, marcarlo para eliminar por seguridad
                            # (esto evita imágenes corruptas que causan problemas)
                            pass
                    except Exception as e_img:
                        # Si cualquier cosa falla al verificar la imagen, ignorarla
                        # print(f"    ⚠️ Imagen con formato inválido detectada, omitiendo verificación")
                        pass
                
                # Eliminar las imágenes marcadas
                for img in imagenes_a_eliminar:
                    try:
                        ws._images.remove(img)
                    except:
                        pass
                
                # Insertar nueva imagen
                img = Image(png_path)
                img.width = 17.7 * 37.8
                img.height = 7 * 37.8
                ws.add_image(img, cell)
                
                print(f"  ✓ {inst_name} → {sheet_name}:{cell}")
                insertados += 1
                
            except Exception as e:
                print(f"  ✗ Error insertando {inst_name}: {e}")
                omitidos += 1
        
        # Guardar Excel UNA SOLA VEZ
        wb.save(excel_path)
        print(f"\n✓ Gráficos insertados: {insertados}")
        if omitidos > 0:
            print(f"⚠️ Gráficos omitidos: {omitidos}")
        
    except Exception as e:
        print(f"✗ Error al abrir/guardar Excel: {e}")
        import traceback
        traceback.print_exc()